"""
Produce output: schedules
"""
from tools.helper import get_cell_specs, cell_fonts, cell_align, cell_border, cell_fill
from tools.xl import write_line, write_col, rectangle_border, rectangle_color, adjust_dimensions
from workflow import schedule

import numpy
from openpyxl import load_workbook, Workbook
from copy import deepcopy
from collections import defaultdict

class Result:
    def __init__(self, specs, db, xdata, varmap, vars):
        self.shout = specs.shout['output']
        # Useful variables
        self.varmap = varmap
        self.nteas, self.nbds = xdata.nteas, xdata.nbds
        self.bands = xdata.bands
        self.sect_names = xdata.sect_names
        self.fcall = self._find_sect_data

        self.daynames = specs.daymarks
        self.dayprds = specs.dayprds
        self.grades = specs.grades
        self.lunch = specs.lunch
        self.shape = specs.shape
        self.colors = specs.colors
        self.outnames = specs.outnames
        self.make_outfile = specs.make_outfile
        
        # Constants for output files
        self.col_header = list(self.daynames)
        mp = max(self.dayprds)
        hdr = list(range(1,mp+1))
        hdr.insert(-2,'M'); hdr.insert(3,'HR'); hdr.insert(1,'T')
        self.row_header = hdr
        self.colAw = 15  # column A width
        
        # undecided pre-process
        schedule.pcs_undecided(vars, xdata.stud_crs_map, xdata.tbd_courses, xdata.decisions, db.students)        
        self.rosters = defaultdict(lambda: {'r':[],'F':0,'M':0})
        # Get roster per section
        schedule.get_rosters(xdata.decisions, db.students, self.rosters)
        self.print_decisions(xdata.decisions, xdata.tbd_courses)
        
        # All schedules
        self.prepare_schedule(db, xdata, vars)
        # Log
        self.print_log(xdata.decisions, specs.shout)


    def print_log(self, decisions, shout):
        """ Report """
        if shout['phase_completion']:
            print("\nPhase 5 complete.")
        if shout['schedule_decisions']:
            print("Decisions for undecided students:")
            for stuid, decs in decisions.items():
                print("  ", stuid)
                for sid, idx in decs:
                    line = "\t{}\t{}"
                    print(line.format(sid,idx))

    def prepare_schedule(self, db, xdata, vars):
        """Prepare and print schedules"""
        pstr = "plan_{}"
        # Bands
        fkey = pstr.format("bands")
        self.pcs_general(fkey, vars, "bands", cohorts=xdata.grades, bands=xdata.bands, sections=db.sections)
        
        # Teacher Schedules
        fkey = pstr.format("faculty")
        self.pcs_general(fkey, vars, "teachers", teachers=db.teachers, tea_sects=xdata.tea_sects)
        
        # Student Schedules
        fkey = pstr.format("students")
        self.pcs_general(fkey, vars, "students", students=db.students, cohorts=xdata.grades, decisions=xdata.decisions)

    def pcs_general(self, fkey, vars, ptype, **kwargs):
        """Create file, call, save"""
        # Create a new blank Workbook object
        wbr = Workbook()
        wbr.title = self.outnames[fkey]
        fname = self.make_outfile(fkey)
        
        if ptype == "teachers":
            self.pcs_teachers(wbr, vars, kwargs['teachers'], kwargs['tea_sects'])
        elif ptype == "students":
            self.pcs_students(wbr, vars, kwargs['students'], kwargs['cohorts'], kwargs['decisions'])
        elif ptype == "bands":
            self.pcs_bands(wbr, vars, kwargs['cohorts'], kwargs['bands'], kwargs['sections'])

        # Save
        wbr.save(fname)
        if self.shout: print("Wrote file: ", fname)

    def pcs_teachers(self, wbr, vars, teachers, tea_sects):
        """ Teachers """
        ptype = "teachers"
        # Schedule skeleton
        schd_week = self.get_skeleton(ptype)
        # Build schedule
        schd_all = self.build_schedule(ptype, schd_week, vars, teachers, tea_sects)
        # Print
        self.print_schedule_teachers(wbr, schd_all, teachers, tea_sects)

    def pcs_students(self, wbr, vars, students, cohorts, decisions):
        """ Students """
        ptype = "students"
        # Each grade separately
        for g in self.grades:
            name = "All Grade {}".format(g)
            page = wbr.create_sheet(name)
            studs = cohorts[g].stud_info
            # Schedule skeleton
            schd_week = self.get_skeleton(ptype)
            # Build schedule
            schd_all = self.build_schedule(ptype, schd_week, vars, students, studs, decisions)
            # Print
            self.print_schedule_students(page, schd_all, studs)

    def pcs_bands(self, wbr, vars, cohorts, bands, sections):
        """ Bands and classes """
        ptype = "bands"
        # Schedule skeleton
        schd_week = self.get_skeleton(ptype)
        # Build schedule
        self.build_schedule(ptype, schd_week, vars)
        # print schedule
        self.print_schedule_bands(wbr, schd_week, cohorts, bands)
        # print entire week
        self.print_week(schd_week, cohorts, bands, sections)


    # ---------------------------------------------------
    # Subroutines
    # ---------------------------------------------------
    def get_skeleton(self, ptype):
        """Get schedule skeleton"""
        if ptype == "bands":
            schd_week = [[[] for p in range(nprds)] for nprds in self.dayprds]
        elif ptype == "teachers":
            schd_week = [[[p+1, "-", ""] for p in range(nprds)] for nprds in self.dayprds]
        elif ptype == "students" or ptype == "teachers":
            schd_week = [[[p+1, "-", ""] for p in range(nprds)] for nprds in self.dayprds]
        return schd_week

    def build_schedule(self, ptype, schd_week, vars, *args):
        if ptype == "bands":
            self.build_week_bands(schd_week, vars)
            return None
        elif ptype == "teachers":
            schd_all = [deepcopy(schd_week) for _ in range(self.nteas)]
            teachers, tea_sects = args
            self.build_week_teachers(schd_all, vars, teachers, tea_sects)
            return schd_all
        elif ptype == "students":
            students, studs, decisions = args
            schd_all = {stud[0]: deepcopy(schd_week) for stud in studs}
            self.build_week_students(schd_all, vars, students, studs, decisions)
            return schd_all

    def build_week_bands(self, schd_week, vars):
        """Weekly schedule for bands"""
        for d, day in enumerate(self.daynames):
            vtp = numpy.rint(vars.tp[day].value)
            vcp = numpy.rint(vars.cp[day].value)
            for s,vc in enumerate(vcp):
                for p in numpy.nonzero(vc)[0]:
                    sid, name = self._find_sect_data(s)
                    schd_week[d][p].append([sid, name, s])

    def build_week_teachers(self, schd_all, vars, teachers, tea_sects):
        """Weekly schedule for teachers"""
        for d, day in enumerate(self.daynames):
            # vars for day
            vtp = numpy.rint(vars.tp[day].value)
            vcp = numpy.rint(vars.cp[day].value)
            # Each teacher
            for t, (tid, idxs) in enumerate(tea_sects.items()):
                vp = vtp[t,:]
                vc = vcp[idxs,:]
                schday = schd_all[t][d]
                schedule.teacher_day_sched(vp, vc, idxs, schday, self.fcall)
                schedule.replace_band_names(schday, self.bands, teachers[tid].sects)

    # For each student
    def build_week_students(self, schd_all, vars, students, studs, decisions):
        """Weekly schedule for students"""

    """
        
    REST OF CODE HAS BEEN OMITTED!!
        
    """
