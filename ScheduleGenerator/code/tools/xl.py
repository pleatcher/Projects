"""
Helper functions for EXCEL files/sheets
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from itertools import islice
import warnings

def read_dframe(fname):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        #df = pd.read_excel(fname, engine='openpyxl')
        # Read first row, get header
        columns = pd.read_excel(fname, engine='openpyxl', nrows=0).columns
        df = pd.read_excel(fname, engine='openpyxl', converters={col: str for col in columns})
    df.columns = [col.lower().replace(" ","_") for col in df]
    df.dropna(inplace=True,how='all')
    df.dropna(inplace=True,axis=1,how='all')
    return (list(df.columns), df.itertuples())

def read_dframe_mul(fname):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df_all = pd.read_excel(fname, engine='openpyxl', sheet_name = None)
    for _, df in df_all.items():
        # Process each sheet
        df.dropna(inplace=True,how='all')
        df.dropna(inplace=True,axis=1,how='all')
    return df_all

def read_rows(sheet):
    """Read all rows from sheet"""
    # Column names
    header = next(sheet.rows)
    # process rows
    rows = []
    for row in islice(sheet.rows, 1, None):
        clean_row = [str(cell.value).strip() for cell in row]
        rows.append(clean_row)
    return (header, rows)

def read_file(fname):
    # Read excel file
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(fname, data_only=True)
        sheet = wb.active
    return read_rows(sheet)

def read_file_multiple(fname, names):
    # Read excel file with multiple sheets
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(fname, data_only=True)
    celldata = []
    for name in names:
        sheet = wb[name]
        celldata.append(read_rows(sheet))    
    return celldata

def write_cell(C, text, cell_font=None, cell_align=None, cell_border=None, cell_fill=None):
    """ Write one cell with props"""
    C.value = text
    if cell_font: C.font = cell_font
    if cell_align: C.alignment = cell_align
    if cell_border: C.border = cell_border
    if cell_fill: C.fill = cell_fill

def write_line(page, line, num, cell_font=None, cell_align=None, cell_border=None, cell_fill=None):
    """ Write one line to excel file"""
    row_num, col_num = num
    if not isinstance(line, list): line = [line]
    for n, text in enumerate(line):
        C = page.cell(row = row_num, column = col_num+n)
        write_cell(C, text, cell_font, cell_align, cell_border, cell_fill)

def write_col(page, col, num, cell_font=None, cell_align=None, cell_border=None, cell_fill=None):
    """ Write one column """
    row_num, col_num = num
    for n, text in enumerate(col):
        C = page.cell(row = row_num+n, column = col_num)
        write_cell(C, text, cell_font, cell_align, cell_border, cell_fill)

def adjust_dimensions(page, w, h):
    for col in page.columns:
        page.column_dimensions[col[0].column_letter].width = w
    for row in range(1, page.max_row+1):
        page.row_dimensions[row].height = h

def rectangle_border(page, cell_border, row_t, row_b, col_l, col_r):
    # top-bottom
    for c in range(col_l,col_r+1):
        page.cell(row=row_t, column=c).border = cell_border['T']
        page.cell(row=row_b, column=c).border = cell_border['B']
    # left-right
    for r in range(row_t,row_b+1):
        page.cell(row=r, column=col_l).border = cell_border['L']
        page.cell(row=r, column=col_r).border = cell_border['R']
    # corners
    page.cell(row=row_t, column=col_l).border = cell_border['LT']
    page.cell(row=row_t, column=col_r).border = cell_border['RT']
    page.cell(row=row_b, column=col_l).border = cell_border['LB']
    page.cell(row=row_b, column=col_r).border = cell_border['RB']

def rectangle_color(page, cell_fill, row_t, row_b, col_l, col_r):
    """Rectangle with solid color"""
    for r in range(row_t,row_b+1):
        for c in range(col_l,col_r+1):
            page.cell(row=r,column=c).fill = cell_fill

"""
def write_cell_bold(page, r, c, v):
    C = page.cell(row=r, column=c, value=v)
    C.font = cell_bold
    C.alignment = cell_align

def write_column_name(page, c):
    colname = chr(65 + c)
    write_cell_bold(page, 2, 2+c, colname)

def write_row_numbers(page, c_rows):
    for r in range(c_rows):
        write_cell_bold(page, 3+r, 1, r+1)
"""
